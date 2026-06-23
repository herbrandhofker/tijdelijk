import openpyxl

excel_path = "Lostock Fabric JV+RdW.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

empty_rows = []
filled_rows = []

for r in range(4, sheet.max_row + 1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 17)]
    # Column B is Quantity (Aantal), Column C is Number (nummer)
    qty = sheet.cell(row=r, column=2).value
    num = sheet.cell(row=r, column=3).value
    if qty is None and num is None:
        empty_rows.append((r, vals))
    else:
        filled_rows.append((r, vals))

print(f"Total rows: {sheet.max_row}")
print(f"Total filled rows: {len(filled_rows)}")
print(f"Total empty/placeholder rows: {len(empty_rows)}")

if empty_rows:
    print(f"Empty rows range: {empty_rows[0][0]} to {empty_rows[-1][0]}")
    print("First 5 empty rows:")
    for r, vals in empty_rows[:5]:
        print(f"  Row {r:03d}: {vals[:8]}")
