import os
import shutil
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def improve_excel():
    excel_path = "/home/herbrand/vragen/Lostock Fabric JV+RdW.xlsx"
    backup_path = "/home/herbrand/vragen/Lostock Fabric JV+RdW_backup.xlsx"
    
    if not os.path.exists(excel_path):
        print(f"Error: {excel_path} not found")
        return
        
    # 1. Back up the original file
    shutil.copyfile(excel_path, backup_path)
    print(f"Backup created at: {backup_path}")
    
    # 2. Load the Workbook
    wb = openpyxl.load_workbook(excel_path)
    
    # Identify the active detailed sheet and rename it
    if "Blad1" in wb.sheetnames:
        ws_details = wb["Blad1"]
        ws_details.title = "Meetstaten Details"
    else:
        # Fallback if already renamed
        ws_details = wb.active
        ws_details.title = "Meetstaten Details"
        
    ws_details.views.sheetView[0].showGridLines = True
    
    # Define styles
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="595959")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    card_title_font = Font(name=font_family, size=9, bold=True, color="595959")
    card_value_font = Font(name=font_family, size=18, bold=True, color="1F497D")
    
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    card_fill = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
    summary_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom = Border(
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="double", color="000000")
    )
    
    # 3. Process the "Meetstaten Details" sheet
    # Correct the Dates in Column R (column 18) based on Item number in Column C (column 3)
    # Range of items is 1 to 333
    # Day 1: items 1-43 -> '09-06-2026'
    # Day 2: items 44-75 -> '10-06-2026'
    # Day 4: items 278-333 -> '12-06-2026'
    
    max_row_before = ws_details.max_row
    print(f"Original max row of details: {max_row_before}")
    
    # Scan rows 4 to 171 to correct dates
    for r in range(4, 172):
        item_val = ws_details.cell(row=r, column=3).value
        if isinstance(item_val, int):
            if 1 <= item_val <= 43:
                ws_details.cell(row=r, column=18, value="09-06-2026")
            elif 44 <= item_val <= 75:
                ws_details.cell(row=r, column=18, value="10-06-2026")
            elif 278 <= item_val <= 333:
                ws_details.cell(row=r, column=18, value="12-06-2026")
                
    # Delete empty rows from 172 to max_row_before
    if max_row_before >= 172:
        ws_details.delete_rows(172, max_row_before - 171)
        print(f"Deleted rows 172 to {max_row_before}")
        
    # Re-verify max row after deletion (should be 171)
    print(f"Max row after deletion: {ws_details.max_row}")
    
    # Format headers of Meetstaten Details (Row 3)
    for c in range(1, 22):
        cell = ws_details.cell(row=3, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Format data rows of Meetstaten Details (Row 4 to 171)
    for idx, r in enumerate(range(4, 172)):
        fill = zebra_fill if idx % 2 == 1 else PatternFill(fill_type=None)
        for c in range(1, 22):
            cell = ws_details.cell(row=r, column=c)
            cell.font = regular_font
            cell.border = thin_border
            if fill.fill_type:
                cell.fill = fill
                
            # Alignments
            if c in [1, 3, 18, 19]: # nummer (code), item_nummer, Datum, Inmeter
                cell.alignment = center_align
            elif c in [2, 4, 5, 9, 10, 11, 12, 13, 14, 15, 16]: # Qty, dimensions, formulas, weights, prices
                cell.alignment = right_align
            else:
                cell.alignment = left_align
                
            # Number Formats
            if c in [2, 3, 12, 14]: # Qty, Item, Price factor, Labels
                cell.number_format = "#,##0"
            elif c in [4, 5, 9, 10, 16]: # dimensions, weights
                cell.number_format = "0.00"
            elif c in [11]: # m2
                cell.number_format = "0.000"
            elif c in [13, 15]: # Prices
                cell.number_format = "#,##0.00"
                
    # Add Total Row in Meetstaten Details at Row 172
    total_row = 172
    ws_details.cell(row=total_row, column=1, value="Totaal").font = bold_font
    ws_details.cell(row=total_row, column=1).alignment = left_align
    
    # Set formulas for totals
    ws_details.cell(row=total_row, column=2, value="=SUM(B4:B171)").font = bold_font # Qty
    ws_details.cell(row=total_row, column=11, value="=SUM(K4:K171)").font = bold_font # m2
    ws_details.cell(row=total_row, column=13, value="=SUM(M4:M171)").font = bold_font # Price excl labels
    ws_details.cell(row=total_row, column=14, value="=SUM(N4:N171)").font = bold_font # Labels
    ws_details.cell(row=total_row, column=15, value="=SUM(O4:O171)").font = bold_font # Total Price incl labels
    ws_details.cell(row=total_row, column=16, value="=SUM(P4:P171)").font = bold_font # kg
    
    # Style total row
    for c in range(1, 22):
        cell = ws_details.cell(row=total_row, column=c)
        cell.border = double_bottom
        cell.fill = summary_fill
        if c in [2, 11, 13, 14, 15, 16]:
            cell.alignment = right_align
            if c in [2, 14]:
                cell.number_format = "#,##0"
            elif c == 11:
                cell.number_format = "0.000"
            elif c in [13, 15]:
                cell.number_format = "#,##0.00"
            elif c == 16:
                cell.number_format = "0.00"
                
    # Adjust column widths
    for col in ws_details.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if type(cell).__name__ == 'MergedCell':
                continue
            if cell.value:
                val_str = str(cell.value)
                if val_str.startswith("="):
                    val_str = "123,456.00" # generic formula placeholder length
                max_len = max(max_len, len(val_str))
        ws_details.column_dimensions[col_letter].width = max(max_len + 3, 10)
    # Override width for notes/specific columns
    ws_details.column_dimensions["A"].width = 16
    ws_details.column_dimensions["G"].width = 18
    ws_details.column_dimensions["Q"].width = 15
    ws_details.column_dimensions["R"].width = 13
    
    # 4. Create Sheet 3: "Meetstaten Pagina's" (Page summaries)
    ws_pages = wb.create_sheet(title="Meetstaten Pagina's")
    ws_pages.views.sheetView[0].showGridLines = True
    
    headers_pages = [
        "Pagina", "Datum", "Dag", "Sectie", "Matrassen (st.)", "Petten (st.)", "Totaal Items", "Foto Pagina's", "Opmerkingen"
    ]
    
    for c_idx, h_text in enumerate(headers_pages, 1):
        cell = ws_pages.cell(row=1, column=c_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    details_pages = [
        # Day 1
        (2, "09-06-2026", "1e dag", "Sectie 1", 156, 72, "5-19", "Geverifieerd"),
        (3, "09-06-2026", "1e dag", "Sectie 1", 108, 56, "5-19", "Geverifieerd"),
        (4, "09-06-2026", "1e dag", "Sectie 1", 43, 6, "5-19", "Geverifieerd; Box 22 marked NV is meegeteld"),
        (20, "09-06-2026", "1e dag", "Sectie 2", 5, 5, "23-35", "Geverifieerd"),
        (21, "09-06-2026", "1e dag", "Sectie 2", 8, 0, "23-35", "Geverifieerd; Geen haken"),
        (22, "09-06-2026", "1e dag", "Sectie 2", 36, 24, "23-35", "Geverifieerd"),
        # Day 2
        (36, "10-06-2026", "2e dag", "Sectie 3", 44, 16, "38-40", "Geverifieerd; Box 50 marked NV is meegeteld"),
        (37, "10-06-2026", "2e dag", "Sectie 3", 18, 4, "38-40", "Geverifieerd"),
        (41, "10-06-2026", "2e dag", "Sectie 4", 28, 18, "42", "Geverifieerd"),
        (43, "10-06-2026", "2e dag", "Sectie 5", 26, 4, "44-46", "Geverifieerd; Incl. doorgestreepte matrassen"),
        (47, "10-06-2026", "2e dag", "Sectie 6", 20, 4, "48", "Geverifieerd"),
        (49, "10-06-2026", "2e dag", "Sectie 7", 12, 1, "50-51", "Geverifieerd"),
        (52, "10-06-2026", "2e dag", "Sectie 8", 21, 0, "60-84", "Geverifieerd"),
        (53, "10-06-2026", "2e dag", "Sectie 8", 21, 0, "60-84", "Geverifieerd"),
        (54, "10-06-2026", "2e dag", "Sectie 8", 11, 0, "60-84", "Geverifieerd"),
        (55, "10-06-2026", "2e dag", "Sectie 8", 15, 0, "60-84", "Geverifieerd"),
        (56, "10-06-2026", "2e dag", "Sectie 8", 16, 2, "60-84", "Geverifieerd"),
        (57, "10-06-2026", "2e dag", "Sectie 8", 20, 0, "60-84", "Geverifieerd; Datum op blad heeft typefout 10-5"),
        (58, "10-06-2026", "2e dag", "Sectie 8", 27, 0, "60-84", "Geverifieerd"),
        (59, "10-06-2026", "2e dag", "Sectie 8", 11, 0, "60-84", "Geverifieerd"),
        # Day 3
        (85, "11-06-2026", "3e dag", "Sectie 9", 14, 0, "90-95", "Geverifieerd"),
        (86, "11-06-2026", "3e dag", "Sectie 9", 33, 0, "90-95", "Geverifieerd"),
        (87, "11-06-2026", "3e dag", "Sectie 9", 81, 1, "90-95", "Geverifieerd; Box 171 doorgestreept en meegeteld"),
        (88, "11-06-2026", "3e dag", "Sectie 9", 9, 0, "90-95", "Geverifieerd"),
        (89, "11-06-2026", "3e dag", "Sectie 9", 7, 0, "90-95", "Geverifieerd"),
        (96, "11-06-2026", "3e dag", "Sectie 10", 48, 0, "97-98", "Geverifieerd"),
        (99, "11-06-2026", "3e dag", "Sectie 11", 24, 9, "104-106", "Geverifieerd; Box 203 doorgestreept en meegeteld"),
        (100, "11-06-2026", "3e dag", "Sectie 11", 51, 0, "104-106", "Geverifieerd"),
        (101, "11-06-2026", "3e dag", "Sectie 11", 42, 0, "104-106", "Geverifieerd"),
        (102, "11-06-2026", "3e dag", "Sectie 11", 48, 0, "104-106", "Geverifieerd"),
        (103, "11-06-2026", "3e dag", "Sectie 11", 45, 0, "104-106", "Geverifieerd"),
        # Day 4
        (107, "12-06-2026", "4e dag", "Sectie 12", 3, 0, "110-111", "Geverifieerd"),
        (108, "12-06-2026", "4e dag", "Sectie 12", 16, 0, "110-111", "Geverifieerd"),
        (109, "12-06-2026", "4e dag", "Sectie 12", 12, 0, "110-111", "Geverifieerd"),
        (112, "12-06-2026", "4e dag", "Sectie 13", 11, 2, "117-131", "Geverifieerd"),
        (113, "12-06-2026", "4e dag", "Sectie 13", 16, 0, "117-131", "Geverifieerd"),
        (114, "12-06-2026", "4e dag", "Sectie 13", 17, 0, "117-131", "Geverifieerd"),
        (115, "12-06-2026", "4e dag", "Sectie 13", 1, 0, "117-131", "Geverifieerd; Slechts 1 matras (met slash /)"),
        (116, "12-06-2026", "4e dag", "Sectie 13", 2, 0, "117-131", "Geverifieerd"),
        (132, "12-06-2026", "4e dag", "Sectie 14", 7, 1, "133-143", "Geverifieerd; Exclusief lijstbestelling"),
        (144, "12-06-2026", "4e dag", "Sectie 15", 35, 17, "148-155", "Geverifieerd"),
        (145, "12-06-2026", "4e dag", "Sectie 15", 11, 2, "148-155", "Geverifieerd"),
        (146, "12-06-2026", "4e dag", "Sectie 15", 10, 0, "148-155", "Geverifieerd"),
        (147, "12-06-2026", "4e dag", "Sectie 15", 3, 0, "148-155", "Geverifieerd")
    ]
    
    for idx, data in enumerate(details_pages):
        row = 2 + idx
        fill = zebra_fill if idx % 2 == 1 else PatternFill(fill_type=None)
        
        page, date, day, sec, matr, pet, photos, notes = data
        
        c_page = ws_pages.cell(row=row, column=1, value=page)
        c_date = ws_pages.cell(row=row, column=2, value=date)
        c_day = ws_pages.cell(row=row, column=3, value=day)
        c_sec = ws_pages.cell(row=row, column=4, value=sec)
        c_matr = ws_pages.cell(row=row, column=5, value=matr)
        c_pet = ws_pages.cell(row=row, column=6, value=pet)
        c_tot = ws_pages.cell(row=row, column=7, value=f"=E{row}+F{row}")
        c_photos = ws_pages.cell(row=row, column=8, value=photos)
        c_notes = ws_pages.cell(row=row, column=9, value=notes)
        
        c_page.alignment = center_align
        c_date.alignment = center_align
        c_day.alignment = center_align
        c_sec.alignment = center_align
        c_matr.alignment = right_align
        c_pet.alignment = right_align
        c_tot.alignment = right_align
        c_photos.alignment = center_align
        c_notes.alignment = left_align
        
        c_page.font = bold_font
        c_tot.font = bold_font
        for c in range(2, 7):
            ws_pages.cell(row=row, column=c).font = regular_font
        c_photos.font = regular_font
        c_notes.font = regular_font
        
        for c in range(1, 10):
            cell = ws_pages.cell(row=row, column=c)
            cell.border = thin_border
            if fill.fill_type:
                cell.fill = fill
            if c in [5, 6, 7]:
                cell.number_format = "#,##0"
                
    # Extra / Lijstbestellingen row (row 46)
    row = 2 + len(details_pages)
    ws_pages.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws_pages.cell(row=row, column=1, value="Extra / Lijstbestellingen (Repetities)").font = bold_font
    ws_pages.cell(row=row, column=1).alignment = left_align
    
    ws_pages.cell(row=row, column=5, value=127).font = bold_font
    ws_pages.cell(row=row, column=5).alignment = right_align
    ws_pages.cell(row=row, column=5).number_format = "#,##0"
    
    ws_pages.cell(row=row, column=6, value=0).font = bold_font
    ws_pages.cell(row=row, column=6).alignment = right_align
    ws_pages.cell(row=row, column=6).number_format = "#,##0"
    
    ws_pages.cell(row=row, column=7, value=f"=E{row}+F{row}").font = bold_font
    ws_pages.cell(row=row, column=7).alignment = right_align
    ws_pages.cell(row=row, column=7).number_format = "#,##0"
    
    ws_pages.cell(row=row, column=9, value="Standard repetities zonder tekening (108 dag 3, 19 dag 4)").font = regular_font
    
    for c in range(1, 10):
        cell = ws_pages.cell(row=row, column=c)
        cell.border = thin_border
        cell.fill = summary_fill
        
    # Total row (row 47)
    row = 3 + len(details_pages)
    ws_pages.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws_pages.cell(row=row, column=1, value="Totaal").font = bold_font
    ws_pages.cell(row=row, column=1).alignment = left_align
    
    ws_pages.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = bold_font
    ws_pages.cell(row=row, column=5).alignment = right_align
    ws_pages.cell(row=row, column=5).number_format = "#,##0"
    
    ws_pages.cell(row=row, column=6, value=f"=SUM(F2:F{row-1})").font = bold_font
    ws_pages.cell(row=row, column=6).alignment = right_align
    ws_pages.cell(row=row, column=6).number_format = "#,##0"
    
    ws_pages.cell(row=row, column=7, value=f"=SUM(G2:G{row-1})").font = bold_font
    ws_pages.cell(row=row, column=7).alignment = right_align
    ws_pages.cell(row=row, column=7).number_format = "#,##0"
    
    for c in range(1, 10):
        cell = ws_pages.cell(row=row, column=c)
        cell.border = double_bottom
        cell.fill = summary_fill
        
    for col in ws_pages.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if type(cell).__name__ == 'MergedCell':
                continue
            if cell.value:
                val_str = str(cell.value)
                if val_str.startswith("="):
                    val_str = "123,456"
                max_len = max(max_len, len(val_str))
        ws_pages.column_dimensions[col_letter].width = max(max_len + 3, 11)
    ws_pages.column_dimensions["I"].width = 50
    
    # 5. Create Sheet 1: "Samenvatting" (Summary Sheet)
    # We will create it as the active sheet or insert it at index 0
    ws_sum = wb.create_sheet(title="Samenvatting", index=0)
    ws_sum.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_sum["A1"] = "Lostock Fabric — Meetstaten Samenvatting"
    ws_sum["A1"].font = title_font
    ws_sum["A2"] = "Productie- en inmeetgegevens isolatiematrassen (SIB Isolatie / Hotrema)"
    ws_sum["A2"].font = subtitle_font
    
    # Metadata Block
    metadata = [
        ("Opdrachtgever:", "Hotrema"),
        ("Project:", "Lostock Fabric"),
        ("Uitvoering:", "Sil-HT700-wol"),
        ("Inmeters:", "JV & RdW (Ron de Wit)"),
        ("Inmeetperiode:", "9 juni 2026 t/m 12 juni 2026"),
        ("Rapportagedatum:", "23 juni 2026")
    ]
    
    for idx, (label, val) in enumerate(metadata):
        row = 4 + idx
        ws_sum.cell(row=row, column=1, value=label).font = bold_font
        ws_sum.cell(row=row, column=2, value=val).font = regular_font
        ws_sum.cell(row=row, column=1).alignment = left_align
        ws_sum.cell(row=row, column=2).alignment = left_align
        
    # KPI Summary Cards (Side by Side)
    # Card 1: Project Totaal
    ws_sum.merge_cells("D4:E4")
    ws_sum["D4"] = "PROJECT TOTAAL (ST.)"
    ws_sum["D4"].font = card_title_font
    ws_sum["D4"].alignment = center_align
    
    ws_sum.merge_cells("D5:E6")
    ws_sum["D5"] = "='Meetstaten Pagina\'s'!G47"
    ws_sum["D5"].font = card_value_font
    ws_sum["D5"].alignment = center_align
    ws_sum["D5"].number_format = "#,##0"
    
    # Card 2: Aantal Matrassen
    ws_sum.merge_cells("G4:H4")
    ws_sum["G4"] = "AANTAL MATRASSEN (ST.)"
    ws_sum["G4"].font = card_title_font
    ws_sum["G4"].alignment = center_align
    
    ws_sum.merge_cells("G5:H6")
    ws_sum["G5"] = "='Meetstaten Pagina\'s'!E47"
    ws_sum["G5"].font = card_value_font
    ws_sum["G5"].alignment = center_align
    ws_sum["G5"].number_format = "#,##0"
    
    # Card 3: Aantal Petten
    ws_sum.merge_cells("J4:K4")
    ws_sum["J4"] = "AANTAL PETTEN"
    ws_sum["J4"].font = card_title_font
    ws_sum["J4"].alignment = center_align
    
    ws_sum.merge_cells("J5:K6")
    ws_sum["J5"] = "='Meetstaten Pagina\'s'!F47"
    ws_sum["J5"].font = card_value_font
    ws_sum["J5"].alignment = center_align
    ws_sum["J5"].number_format = "#,##0"
    
    # Card 4: Database Details (actually detailed items in database)
    ws_sum.merge_cells("M4:N4")
    ws_sum["M4"] = "DATABASE DETAILS (ST.)"
    ws_sum["M4"].font = card_title_font
    ws_sum["M4"].alignment = center_align
    
    ws_sum.merge_cells("M5:N6")
    ws_sum["M5"] = "='Meetstaten Details'!B172"
    ws_sum["M5"].font = card_value_font
    ws_sum["M5"].alignment = center_align
    ws_sum["M5"].number_format = "#,##0"
    
    # Style KPI cards
    for col_range in ["D4:E6", "G4:H6", "J4:K6", "M4:N6"]:
        start_col, start_row = col_range.split(":")[0][0], int(col_range.split(":")[0][1])
        end_col, end_row = col_range.split(":")[1][0], int(col_range.split(":")[1][1])
        
        start_c_idx = ord(start_col) - ord('A') + 1
        end_c_idx = ord(end_col) - ord('A') + 1
        
        for r in range(start_row, end_row + 1):
            for c in range(start_c_idx, end_c_idx + 1):
                cell = ws_sum.cell(row=r, column=c)
                cell.fill = card_fill
                cell.border = thin_border
                
    # Daily production table headers
    ws_sum["A12"] = "Dag"
    ws_sum["B12"] = "Datum"
    ws_sum["C12"] = "Project Totaal (st.)"
    ws_sum["D12"] = "Database Details (st.)"
    ws_sum["E12"] = "m2 Details (m2)"
    ws_sum["F12"] = "Gewicht Details (kg)"
    ws_sum["G12"] = "Waarde Details (€)"
    
    for col in ["A12", "B12", "C12", "D12", "E12", "F12", "G12"]:
        cell = ws_sum[col]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Table data
    # (Day, Date, ProjectTotal)
    prod_data = [
        ("1e dag", "09-06-2026", 519),
        ("2e dag", "10-06-2026", 339),
        ("3e dag", "11-06-2026", 520),
        ("4e dag", "12-06-2026", 185)
    ]
    
    for idx, (day, date, total) in enumerate(prod_data):
        row = 13 + idx
        ws_sum.cell(row=row, column=1, value=day).font = bold_font
        ws_sum.cell(row=row, column=2, value=date).font = regular_font
        ws_sum.cell(row=row, column=3, value=total).font = regular_font
        
        # Formulas to link to details
        # B = Qty, R = Date string, K = m2, P = Weight, O = Price
        ws_sum.cell(row=row, column=4, value=f"=SUMIFS('Meetstaten Details'!B$4:B$171, 'Meetstaten Details'!R$4:R$171, \"{date}\")").font = regular_font
        ws_sum.cell(row=row, column=5, value=f"=SUMIFS('Meetstaten Details'!K$4:K$171, 'Meetstaten Details'!R$4:R$171, \"{date}\")").font = regular_font
        ws_sum.cell(row=row, column=6, value=f"=SUMIFS('Meetstaten Details'!P$4:P$171, 'Meetstaten Details'!R$4:R$171, \"{date}\")").font = regular_font
        ws_sum.cell(row=row, column=7, value=f"=SUMIFS('Meetstaten Details'!O$4:O$171, 'Meetstaten Details'!R$4:R$171, \"{date}\")").font = regular_font
        
        ws_sum.cell(row=row, column=1).alignment = center_align
        ws_sum.cell(row=row, column=2).alignment = center_align
        for c in range(3, 8):
            ws_sum.cell(row=row, column=c).alignment = right_align
            if c in [3, 4]:
                ws_sum.cell(row=row, column=c).number_format = "#,##0"
            elif c == 5:
                ws_sum.cell(row=row, column=c).number_format = "0.000"
            elif c == 6:
                ws_sum.cell(row=row, column=c).number_format = "0.00"
            elif c == 7:
                ws_sum.cell(row=row, column=c).number_format = "#,##0.00"
                
        for c in range(1, 8):
            ws_sum.cell(row=row, column=c).border = thin_border
            
    # Total Row (row 17)
    ws_sum["A17"] = "Totaal"
    ws_sum["A17"].font = bold_font
    ws_sum["A17"].alignment = center_align
    ws_sum["B17"] = ""
    
    ws_sum["C17"] = "=SUM(C13:C16)"
    ws_sum["D17"] = "=SUM(D13:D16)"
    ws_sum["E17"] = "=SUM(E13:E16)"
    ws_sum["F17"] = "=SUM(F13:F16)"
    ws_sum["G17"] = "=SUM(G13:G16)"
    
    for c in range(3, 8):
        cell = ws_sum.cell(row=17, column=c)
        cell.font = bold_font
        cell.alignment = right_align
        if c in [3, 4]:
            cell.number_format = "#,##0"
        elif c == 5:
            cell.number_format = "0.000"
        elif c == 6:
            cell.number_format = "0.00"
        elif c == 7:
            cell.number_format = "#,##0.00"
            
    for c in range(1, 8):
        cell = ws_sum.cell(row=17, column=c)
        cell.border = double_bottom
        cell.fill = summary_fill
        
    for col in ws_sum.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if type(cell).__name__ == 'MergedCell':
                continue
            if cell.value:
                val_str = str(cell.value)
                if val_str.startswith("="):
                    val_str = "123,456.00"
                max_len = max(max_len, len(val_str))
        ws_sum.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    ws_sum.column_dimensions["A"].width = 15
    ws_sum.column_dimensions["B"].width = 15
    ws_sum.column_dimensions["C"].width = 20
    ws_sum.column_dimensions["D"].width = 22
    ws_sum.column_dimensions["E"].width = 18
    ws_sum.column_dimensions["F"].width = 20
    ws_sum.column_dimensions["G"].width = 20
    
    # Save the Workbook
    wb.save(excel_path)
    print(f"Excel workbook updated successfully at {excel_path}")

if __name__ == "__main__":
    improve_excel()
