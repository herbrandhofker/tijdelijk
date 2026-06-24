import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_excel():
    wb = Workbook()
    
    # -------------------------------------------------------------
    # STYLES DEFINITION
    # -------------------------------------------------------------
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="595959")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    card_title_font = Font(name=font_family, size=9, bold=True, color="595959")
    card_value_font = Font(name=font_family, size=18, bold=True, color="1F497D")
    
    # Fills
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # Very Light Blue-Gray
    card_fill = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid") # Light Blue Accent
    summary_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # Total summary fill
    
    # Alignments
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    double_bottom = Border(
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="double", color="000000")
    )
    
    # -------------------------------------------------------------
    # SHEET 1: SAMENVATTING (SUMMARY)
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Samenvatting"
    ws1.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws1["A1"] = "Lostock Fabric — Meetstaten Samenvatting"
    ws1["A1"].font = title_font
    ws1["A2"] = "Productie- en inmeetgegevens isolatiematrassen (SIB Isolatie / Hotrema)"
    ws1["A2"].font = subtitle_font
    
    # Metadata Block
    metadata = [
        ("Opdrachtgever:", "Hotrema"),
        ("Project:", "Lostock Fabric"),
        ("Uitvoering:", "Sil-HT700-wol"),
        ("Inmeters:", "JV & RdW (Ron de Wit)"),
        ("Inmeetperiode:", "9 juni 2026 t/m 12 juni 2026"),
        ("Rapportagedatum:", "23 juni 2026")
    ]
    
    for idx, (label, val) in enumerate(metadata):
        row = 4 + idx
        ws1.cell(row=row, column=1, value=label).font = bold_font
        ws1.cell(row=row, column=2, value=val).font = regular_font
        ws1.cell(row=row, column=1).alignment = left_align
        ws1.cell(row=row, column=2).alignment = left_align
        
    # KPI Summary Cards (Side by Side)
    # Card 1: Totaal Productie
    ws1.merge_cells("D4:E4")
    ws1["D4"] = "TOTAAL PRODUCTIE"
    ws1["D4"].font = card_title_font
    ws1["D4"].alignment = center_align
    
    ws1.merge_cells("D5:E6")
    ws1["D5"] = "=SUM(C14:C17)"
    ws1["D5"].font = card_value_font
    ws1["D5"].alignment = center_align
    ws1["D5"].number_format = "#,##0"
    
    # Card 2: Aantal Matrassen
    ws1.merge_cells("G4:H4")
    ws1["G4"] = "AANTAL MATRASSEN (ST.)"
    ws1["G4"].font = card_title_font
    ws1["G4"].alignment = center_align
    
    ws1.merge_cells("G5:H6")
    ws1["G5"] = 1456
    ws1["G5"].font = card_value_font
    ws1["G5"].alignment = center_align
    ws1["G5"].number_format = "#,##0"
    
    # Card 3: Aantal Petten
    ws1.merge_cells("J4:K4")
    ws1["J4"] = "AANTAL PETTEN"
    ws1["J4"].font = card_title_font
    ws1["J4"].alignment = center_align
    
    ws1.merge_cells("J5:K6")
    ws1["J5"] = 107
    ws1["J5"].font = card_value_font
    ws1["J5"].alignment = center_align
    ws1["J5"].number_format = "#,##0"
    
    # Apply styling to KPI cards
    for col_range in ["D4:E6", "G4:H6", "J4:K6"]:
        start_col, start_row = col_range.split(":")[0][0], int(col_range.split(":")[0][1])
        end_col, end_row = col_range.split(":")[1][0], int(col_range.split(":")[1][1])
        
        start_c_idx = ord(start_col) - ord('A') + 1
        end_c_idx = ord(end_col) - ord('A') + 1
        
        for r in range(start_row, end_row + 1):
            for c in range(start_c_idx, end_c_idx + 1):
                cell = ws1.cell(row=r, column=c)
                cell.fill = card_fill
                cell.border = thin_border
                
    # Production table headers
    ws1["A12"] = "Dag"
    ws1["B12"] = "Datum"
    ws1["C12"] = "Totaal Items"
    
    for col in ["A12", "B12", "C12"]:
        cell = ws1[col]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Production Table Data
    prod_data = [
        ("1e dag", "9-6-2026", 519),
        ("2e dag", "10-6-2026", 339),
        ("3e dag", "11-6-2026", 520),
        ("4e dag", "12-6-2026", 185)
    ]
    
    for idx, (day, date, total) in enumerate(prod_data):
        row = 13 + idx
        ws1.cell(row=row, column=1, value=day).font = bold_font
        ws1.cell(row=row, column=2, value=date).font = regular_font
        ws1.cell(row=row, column=3, value=total).font = regular_font
        
        ws1.cell(row=row, column=1).alignment = center_align
        ws1.cell(row=row, column=2).alignment = center_align
        ws1.cell(row=row, column=3).alignment = right_align
        ws1.cell(row=row, column=3).number_format = "#,##0"
        
        for c in range(1, 4):
            ws1.cell(row=row, column=c).border = thin_border
            
    # Production Table Total Row
    ws1["A17"] = "Totaal"
    ws1["A17"].font = bold_font
    ws1["A17"].alignment = center_align
    ws1["B17"] = ""
    ws1["C17"] = "=SUM(C13:C16)"
    ws1["C17"].font = bold_font
    ws1["C17"].alignment = right_align
    ws1["C17"].number_format = "#,##0"
    
    for c in range(1, 4):
        cell = ws1.cell(row=17, column=c)
        cell.border = double_bottom
        cell.fill = summary_fill

    # -------------------------------------------------------------
    # SHEET 2: MEETSTATEN DETAILS
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="Meetstaten Details")
    ws2.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Pagina", "Datum", "Dag", "Sectie", "Matrassen (st.)", "Petten (st.)", "Totaal Items", "Foto Pagina's", "Opmerkingen"
    ]
    
    for c_idx, h_text in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c_idx, value=h_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Detail Data
    # List: (Page, Date, Day, Section, Mattresses, Petten, PhotoRange, Notes)
    details = [
        # Day 1
        (2, "9-6-2026", "1e dag", "Sectie 1", 156, 72, "5-19", "Geverifieerd"),
        (3, "9-6-2026", "1e dag", "Sectie 1", 108, 56, "5-19", "Geverifieerd"),
        (4, "9-6-2026", "1e dag", "Sectie 1", 43, 6, "5-19", "Geverifieerd; Box 22 marked NV is meegeteld"),
        (20, "9-6-2026", "1e dag", "Sectie 2", 5, 5, "23-35", "Geverifieerd"),
        (21, "9-6-2026", "1e dag", "Sectie 2", 8, 0, "23-35", "Geverifieerd; Geen haken"),
        (22, "9-6-2026", "1e dag", "Sectie 2", 36, 24, "23-35", "Geverifieerd"),
        # Day 2
        (36, "10-6-2026", "2e dag", "Sectie 3", 44, 16, "38-40", "Geverifieerd; Box 50 marked NV is meegeteld"),
        (37, "10-6-2026", "2e dag", "Sectie 3", 18, 4, "38-40", "Geverifieerd"),
        (41, "10-6-2026", "2e dag", "Sectie 4", 28, 18, "42", "Geverifieerd"),
        (43, "10-6-2026", "2e dag", "Sectie 5", 26, 4, "44-46", "Geverifieerd; Incl. doorgestreepte matrassen"),
        (47, "10-6-2026", "2e dag", "Sectie 6", 20, 4, "48", "Geverifieerd"),
        (49, "10-6-2026", "2e dag", "Sectie 7", 12, 1, "50-51", "Geverifieerd"),
        (52, "10-6-2026", "2e dag", "Sectie 8", 21, 0, "60-84", "Geverifieerd"),
        (53, "10-6-2026", "2e dag", "Sectie 8", 21, 0, "60-84", "Geverifieerd"),
        (54, "10-6-2026", "2e dag", "Sectie 8", 11, 0, "60-84", "Geverifieerd"),
        (55, "10-6-2026", "2e dag", "Sectie 8", 15, 0, "60-84", "Geverifieerd"),
        (56, "10-6-2026", "2e dag", "Sectie 8", 16, 2, "60-84", "Geverifieerd"),
        (57, "10-6-2026", "2e dag", "Sectie 8", 20, 0, "60-84", "Geverifieerd; Datum op blad heeft typefout 10-5"),
        (58, "10-6-2026", "2e dag", "Sectie 8", 27, 0, "60-84", "Geverifieerd"),
        (59, "10-6-2026", "2e dag", "Sectie 8", 11, 0, "60-84", "Geverifieerd"),
        # Day 3
        (85, "11-6-2026", "3e dag", "Sectie 9", 14, 0, "90-95", "Geverifieerd"),
        (86, "11-6-2026", "3e dag", "Sectie 9", 33, 0, "90-95", "Geverifieerd"),
        (87, "11-6-2026", "3e dag", "Sectie 9", 81, 1, "90-95", "Geverifieerd; Box 171 doorgestreept en meegeteld"),
        (88, "11-6-2026", "3e dag", "Sectie 9", 9, 0, "90-95", "Geverifieerd"),
        (89, "11-6-2026", "3e dag", "Sectie 9", 7, 0, "90-95", "Geverifieerd"),
        (96, "11-6-2026", "3e dag", "Sectie 10", 48, 0, "97-98", "Geverifieerd"),
        (99, "11-6-2026", "3e dag", "Sectie 11", 24, 9, "104-106", "Geverifieerd; Box 203 doorgestreept en meegeteld"),
        (100, "11-6-2026", "3e dag", "Sectie 11", 51, 0, "104-106", "Geverifieerd"),
        (101, "11-6-2026", "3e dag", "Sectie 11", 42, 0, "104-106", "Geverifieerd"),
        (102, "11-6-2026", "3e dag", "Sectie 11", 48, 0, "104-106", "Geverifieerd"),
        (103, "11-6-2026", "3e dag", "Sectie 11", 45, 0, "104-106", "Geverifieerd"),
        # Day 4
        (107, "12-6-2026", "4e dag", "Sectie 12", 3, 0, "110-111", "Geverifieerd"),
        (108, "12-6-2026", "4e dag", "Sectie 12", 16, 0, "110-111", "Geverifieerd"),
        (109, "12-6-2026", "4e dag", "Sectie 12", 12, 0, "110-111", "Geverifieerd"),
        (112, "12-6-2026", "4e dag", "Sectie 13", 11, 2, "117-131", "Geverifieerd"),
        (113, "12-6-2026", "4e dag", "Sectie 13", 16, 0, "117-131", "Geverifieerd"),
        (114, "12-6-2026", "4e dag", "Sectie 13", 17, 0, "117-131", "Geverifieerd"),
        (115, "12-6-2026", "4e dag", "Sectie 13", 1, 0, "117-131", "Geverifieerd; Slechts 1 matras (met slash /)"),
        (116, "12-6-2026", "4e dag", "Sectie 13", 2, 0, "117-131", "Geverifieerd"),
        (132, "12-6-2026", "4e dag", "Sectie 14", 7, 1, "133-143", "Geverifieerd; Exclusief lijstbestelling"),
        (144, "12-6-2026", "4e dag", "Sectie 15", 35, 17, "148-155", "Geverifieerd"),
        (145, "12-6-2026", "4e dag", "Sectie 15", 11, 2, "148-155", "Geverifieerd"),
        (146, "12-6-2026", "4e dag", "Sectie 15", 10, 0, "148-155", "Geverifieerd"),
        (147, "12-6-2026", "4e dag", "Sectie 15", 3, 0, "148-155", "Geverifieerd")
    ]
    
    for idx, data in enumerate(details):
        row = 2 + idx
        # Zebra striping
        fill = zebra_fill if idx % 2 == 1 else PatternFill(fill_type=None)
        
        # Unpack
        page, date, day, sec, matr, pet, photos, notes = data
        
        # Cells
        c_page = ws2.cell(row=row, column=1, value=page)
        c_date = ws2.cell(row=row, column=2, value=date)
        c_day = ws2.cell(row=row, column=3, value=day)
        c_sec = ws2.cell(row=row, column=4, value=sec)
        c_matr = ws2.cell(row=row, column=5, value=matr)
        c_pet = ws2.cell(row=row, column=6, value=pet)
        
        # Excel Formula for Total Items
        c_tot = ws2.cell(row=row, column=7, value=f"=E{row}+F{row}")
        
        c_photos = ws2.cell(row=row, column=8, value=photos)
        c_notes = ws2.cell(row=row, column=9, value=notes)
        
        # Alignments & Fonts
        c_page.alignment = center_align
        c_date.alignment = center_align
        c_day.alignment = center_align
        c_sec.alignment = center_align
        c_matr.alignment = right_align
        c_pet.alignment = right_align
        c_tot.alignment = right_align
        c_photos.alignment = center_align
        c_notes.alignment = left_align
        
        c_page.font = bold_font
        c_tot.font = bold_font
        for c in range(2, 7):
            ws2.cell(row=row, column=c).font = regular_font
        c_photos.font = regular_font
        c_notes.font = regular_font
        
        # Apply border & fill
        for c in range(1, 10):
            cell = ws2.cell(row=row, column=c)
            cell.border = thin_border
            if fill.fill_type:
                cell.fill = fill
            # Number formats
            if c in [5, 6, 7]:
                cell.number_format = "#,##0"

    # Add row for Extra/Lijstbestellingen
    row = 2 + len(details)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws2.cell(row=row, column=1, value="Extra / Lijstbestellingen (Repetities)").font = bold_font
    ws2.cell(row=row, column=1).alignment = left_align
    
    # 127 extra items total (108 for Day 3, 19 for Day 4)
    ws2.cell(row=row, column=5, value=127).font = bold_font
    ws2.cell(row=row, column=5).alignment = right_align
    ws2.cell(row=row, column=5).number_format = "#,##0"
    
    ws2.cell(row=row, column=6, value=0).font = bold_font
    ws2.cell(row=row, column=6).alignment = right_align
    ws2.cell(row=row, column=6).number_format = "#,##0"
    
    ws2.cell(row=row, column=7, value=f"=E{row}+F{row}").font = bold_font
    ws2.cell(row=row, column=7).alignment = right_align
    ws2.cell(row=row, column=7).number_format = "#,##0"
    
    ws2.cell(row=row, column=9, value="Standard repetities zonder tekening (108 dag 3, 19 dag 4)").font = regular_font
    
    for c in range(1, 10):
        cell = ws2.cell(row=row, column=c)
        cell.border = thin_border
        cell.fill = summary_fill

    # Total Row
    row = 3 + len(details)
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws2.cell(row=row, column=1, value="Totaal").font = bold_font
    ws2.cell(row=row, column=1).alignment = left_align
    
    ws2.cell(row=row, column=5, value=f"=SUM(E2:E{row-1})").font = bold_font
    ws2.cell(row=row, column=5).alignment = right_align
    ws2.cell(row=row, column=5).number_format = "#,##0"
    
    ws2.cell(row=row, column=6, value=f"=SUM(F2:F{row-1})").font = bold_font
    ws2.cell(row=row, column=6).alignment = right_align
    ws2.cell(row=row, column=6).number_format = "#,##0"
    
    ws2.cell(row=row, column=7, value=f"=SUM(G2:G{row-1})").font = bold_font
    ws2.cell(row=row, column=7).alignment = right_align
    ws2.cell(row=row, column=7).number_format = "#,##0"
    
    for c in range(1, 10):
        cell = ws2.cell(row=row, column=c)
        cell.border = double_bottom
        cell.fill = summary_fill
        
    # Auto-adjust column widths for both sheets
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            
            # Skip checking merged cells sizes to prevent extreme widths
            for cell in col:
                # Avoid checking merged values
                if type(cell).__name__ == 'MergedCell':
                    continue
                if cell.value:
                    val_str = str(cell.value)
                    if val_str.startswith("="):
                        val_str = "123,456" # placeholder length for formulas
                    max_len = max(max_len, len(val_str))
            
            ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    # Specific adjustment for notes column
    ws2.column_dimensions["I"].width = 50
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 15
    ws1.column_dimensions["C"].width = 15
    
    # Save Workbook
    output_path = "/home/herbrand/vragen/Lostock_Fabric_Metingen.xlsx"
    wb.save(output_path)
    print(f"Excel workbook created successfully at {output_path}")

if __name__ == "__main__":
    create_excel()
