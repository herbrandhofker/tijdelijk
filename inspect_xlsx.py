import openpyxl

excel_path = "Lostock Fabric JV+RdW.xlsx"
wb = openpyxl.load_workbook(excel_path)
print("Sheet names:", wb.sheetnames)
sheet = wb.active
print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

# Print first 20 rows
print("\n--- FIRST 20 ROWS ---")
for r in range(1, 21):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 17)]
    print(f"Row {r:02d}: {vals}")

# Print a few rows where matras number is in the 70-85 range
print("\n--- ROWS AROUND 70-85 ---")
for r in range(70, 95):
    if r <= sheet.max_row:
        vals = [sheet.cell(row=r, column=c).value for c in range(1, 17)]
        print(f"Row {r:03d}: {vals}")

# Print last 20 rows that have any content
print("\n--- LAST 20 ROWS ---")
found_count = 0
for r in range(sheet.max_row, 0, -1):
    vals = [sheet.cell(row=r, column=c).value for c in range(1, 17)]
    if any(v is not None for v in vals):
        print(f"Row {r:03d}: {vals}")
        found_count += 1
        if found_count >= 20:
            break
