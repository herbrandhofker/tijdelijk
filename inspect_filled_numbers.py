import openpyxl

excel_path = "Lostock Fabric JV+RdW.xlsx"
wb = openpyxl.load_workbook(excel_path)
sheet = wb.active

print("Filled rows:")
for r in range(4, 175):
    qty = sheet.cell(row=r, column=2).value
    num = sheet.cell(row=r, column=3).value
    if qty is not None or num is not None:
        print(f"Row {r:03d}: Number={num}, Qty={qty}, Circumference={sheet.cell(row=r, column=4).value}, Height={sheet.cell(row=r, column=5).value}")
